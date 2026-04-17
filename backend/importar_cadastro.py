"""Importa WOs da aba Cadastro do Excel para a tabela projetos."""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from app.db.database import SessionLocal
from app.models.models import Projeto

ARQUIVO = r'C:\Users\helia\Downloads\Contas a Receber DEV.xlsx'

wb = openpyxl.load_workbook(ARQUIVO, data_only=True)
ws = wb['Cadastro']

# Cabecalhos na linha 4, dados a partir da linha 5
# Colunas: B=WO, C=CLIENTE, D=PLATAFORMA, E=FOCAL, F=COORD FOCAL,
#          G=TIPO DE SERVICO, H=EXTERIOR COM ISS, I=VL.DIARIA,
#          J=VL.DIARIA LOCACAO, K=VL.OUTROS

db = SessionLocal()
try:
    wos_existentes = {p.wo for p in db.query(Projeto.wo).all()}
    inseridos = 0
    ignorados = 0
    erros = []

    for row in range(5, ws.max_row + 1):
        wo_val = ws.cell(row, 2).value
        if not wo_val:
            continue
        try:
            wo = int(wo_val)
        except (ValueError, TypeError):
            continue

        if wo in wos_existentes:
            ignorados += 1
            continue

        cliente    = ws.cell(row, 3).value
        plataforma = ws.cell(row, 4).value
        # E=FOCAL e F=COORD FOCAL deixados vazios conforme instrucao
        tipo_raw   = ws.cell(row, 7).value
        ext_raw    = ws.cell(row, 8).value
        vl_diaria  = ws.cell(row, 9).value
        vl_diaria_loc = ws.cell(row, 10).value
        vl_outros  = ws.cell(row, 11).value

        tipo_servico = str(tipo_raw).strip() if tipo_raw else None
        exterior_com_iss = str(ext_raw).strip().upper() == 'X' if ext_raw else False

        try:
            db.add(Projeto(
                wo=wo,
                cliente=str(cliente).strip() if cliente else None,
                plataforma=str(plataforma).strip() if plataforma else None,
                coordenador=None,   # COORD FOCAL - deixado vazio
                tipo_servico=tipo_servico,
                exterior_com_iss=exterior_com_iss,
                vl_diaria=float(vl_diaria) if vl_diaria else None,
                vl_diaria_locacao=float(vl_diaria_loc) if vl_diaria_loc else None,
                vl_outros=float(vl_outros) if vl_outros else None,
            ))
            wos_existentes.add(wo)
            inseridos += 1
        except Exception as e:
            erros.append(f'WO {wo}: {e}')

    db.commit()
    print(f'OK: {inseridos} projetos inseridos, {ignorados} ja existiam.')
    if erros:
        print(f'Erros ({len(erros)}):')
        for e in erros[:10]:
            print(f'  {e}')
finally:
    db.close()
